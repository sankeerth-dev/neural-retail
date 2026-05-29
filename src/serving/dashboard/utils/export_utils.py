"""NeuralRetail Dashboard — Export Utilities.

Day 16 — NeuralRetail AMX-DS-2026-04
Produces Excel workbooks (openpyxl), PDF reports (ReportLab), and
UTF-8 BOM CSV files for Streamlit download buttons.
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Brand colours as RGB tuples (for ReportLab)
# ---------------------------------------------------------------------------
_PRIMARY_HEX = "E84E1B"
_SECONDARY_HEX = "F7941D"
_ACCENT_HEX = "FBBA13"


class ExportUtils:
    """Utility class for dashboard data exports.

    Provides three export formats:
    - Excel (openpyxl) with conditional formatting on MAPE column.
    - PDF (ReportLab) with NeuralRetail header, metrics table, and chart images.
    - CSV (UTF-8 BOM) for Excel-compatible CRM exports.

    All methods return ``bytes`` objects suitable for ``st.download_button``.
    """

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    @staticmethod
    def forecast_to_excel(
        forecast_df: pd.DataFrame,
        mape_summary_df: pd.DataFrame,
        output_path: str | Path | None = None,
    ) -> bytes:
        """Export forecast data + MAPE summary to a styled Excel workbook.

        Sheet 1 — ``Forecast``: full forecast DataFrame with columns
        [date, actual, p10, p50, p90, is_forecast].

        Sheet 2 — ``MAPE Summary``: per-SKU MAPE with conditional colour
        coding (green ≤ 10%, amber 10-15%, red > 15%).

        Args:
            forecast_df: Forecast DataFrame (typically from demand models).
            mape_summary_df: MAPE leaderboard DataFrame with a ``mape`` column.
            output_path: Optional file path to save the workbook to disk.
                If None, only bytes are returned.

        Returns:
            Excel workbook bytes for ``st.download_button``.

        Raises:
            ImportError: If ``openpyxl`` is not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            logger.error("openpyxl not installed: %s", exc)
            raise

        buffer = io.BytesIO()
        wb = openpyxl.Workbook()

        # ── Shared style helpers ───────────────────────────────────────
        def _header_style(cell: Any) -> None:
            cell.fill = PatternFill("solid", fgColor=_PRIMARY_HEX)
            cell.font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        def _thin_border() -> Border:
            side = Side(style="thin", color="DDDDDD")
            return Border(left=side, right=side, top=side, bottom=side)

        def _auto_width(ws: Any, col_idx: int, value: str) -> None:
            col_letter = get_column_letter(col_idx)
            current = ws.column_dimensions[col_letter].width
            ws.column_dimensions[col_letter].width = max(current, len(str(value)) + 4)

        # ── Sheet 1: Forecast ──────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Forecast"
        ws1.freeze_panes = "A2"

        # Prepare data
        export_df = forecast_df.copy()
        if "date" in export_df.columns and hasattr(export_df["date"], "dt"):
            export_df["date"] = export_df["date"].dt.strftime("%Y-%m-%d")
        elif "date" in export_df.columns:
            export_df["date"] = export_df["date"].astype(str)

        # Header row
        for col_idx, col_name in enumerate(export_df.columns, 1):
            cell = ws1.cell(row=1, column=col_idx, value=col_name)
            _header_style(cell)
            _auto_width(ws1, col_idx, col_name)

        # Data rows
        even_fill = PatternFill("solid", fgColor="F9F9F9")
        for row_idx, row in enumerate(export_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="center")
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        # ── Sheet 2: MAPE Summary ──────────────────────────────────────
        ws2 = wb.create_sheet("MAPE Summary")
        ws2.freeze_panes = "A2"

        green_fill = PatternFill("solid", fgColor="DCFCE7")
        amber_fill = PatternFill("solid", fgColor="FEF3C7")
        red_fill = PatternFill("solid", fgColor="FEE2E2")

        for col_idx, col_name in enumerate(mape_summary_df.columns, 1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            _header_style(cell)
            _auto_width(ws2, col_idx, col_name)

        mape_col_idx: int | None = None
        if "mape" in mape_summary_df.columns:
            mape_col_idx = list(mape_summary_df.columns).index("mape") + 1

        for row_idx, row in enumerate(mape_summary_df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = _thin_border()
                cell.alignment = Alignment(horizontal="center")
                if col_idx == mape_col_idx and isinstance(value, (int, float)):
                    if value <= 10:
                        cell.fill = green_fill
                    elif value <= 15:
                        cell.fill = amber_fill
                    else:
                        cell.fill = red_fill
                elif row_idx % 2 == 0:
                    cell.fill = even_fill

        # Metadata sheet
        ws3 = wb.create_sheet("Metadata")
        ws3["A1"] = "NeuralRetail Intelligence Platform"
        ws3["A1"].font = Font(bold=True, size=14, color=_PRIMARY_HEX)
        ws3["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
        ws3["A3"] = "AMX-DS-2026-04 — CONFIDENTIAL"
        ws3["A4"] = f"Forecast rows: {len(forecast_df)}"
        ws3["A5"] = f"MAPE summary SKUs: {len(mape_summary_df)}"

        if output_path is not None:
            wb.save(str(output_path))
            logger.info("Excel workbook saved to %s", output_path)

        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # PDF export
    # ------------------------------------------------------------------

    @staticmethod
    def kpi_to_pdf(
        metrics_dict: dict[str, Any],
        chart_images: list[bytes],
        output_path: str | Path | None = None,
        title: str = "Executive KPI Report",
    ) -> bytes:
        """Generate a branded PDF report using ReportLab.

        Structure:
        - NeuralRetail orange-gradient header bar.
        - Report title and generation timestamp.
        - Metrics table (KPI | Value | Target | Status).
        - Up to 3 embedded chart PNG images.
        - Footer with project identifier and page number.

        Args:
            metrics_dict: Dict of ``{kpi_name: {value, target, unit, delta_pct}}``.
            chart_images: List of PNG image bytes (up to 3 charts embedded).
            output_path: Optional file path to also save to disk.
            title: Report title string.

        Returns:
            PDF content bytes for ``st.download_button``.

        Raises:
            ImportError: If ``reportlab`` is not installed.
        """
        try:
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                Image,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
                HRFlowable,
            )
        except ImportError as exc:
            logger.error("reportlab not installed: %s", exc)
            raise

        buffer = io.BytesIO()
        PAGE_W, PAGE_H = A4
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
            leftMargin=2 * cm,
            rightMargin=2 * cm,
        )
        styles = getSampleStyleSheet()
        story = []

        # ── Header ────────────────────────────────────────────────────
        brand_style = ParagraphStyle(
            "brand",
            parent=styles["Title"],
            fontSize=22,
            spaceAfter=4,
        )
        story.append(
            Paragraph(
                f'<font color="#{_PRIMARY_HEX}"><b>Neural</b></font>'
                f'<font color="#{_SECONDARY_HEX}"><b>Retail</b></font>'
                f' <font size="14" color="#888888">Intelligence Platform</font>',
                brand_style,
            )
        )
        story.append(
            HRFlowable(
                width="100%",
                thickness=4,
                color=rl_colors.HexColor(f"#{_PRIMARY_HEX}"),
                spaceAfter=8,
            )
        )
        story.append(
            Paragraph(
                f"<b>{title}</b>",
                styles["Heading2"],
            )
        )
        story.append(
            Paragraph(
                f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} &nbsp;&nbsp;|&nbsp;&nbsp; AMX-DS-2026-04",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 0.75 * cm))

        # ── Metrics table ──────────────────────────────────────────────
        table_data = [["KPI", "Current Value", "Target", "Status"]]
        kpi_human_names = {
            "demand_mape": "Demand MAPE",
            "churn_auc": "Churn AUC-ROC",
            "stockout_rate": "Stockout Rate",
            "revenue_uplift": "Revenue Uplift",
        }
        for key, kpi in metrics_dict.items():
            val = kpi.get("value", "N/A")
            tgt = kpi.get("target", "N/A")
            unit = kpi.get("unit", "")
            label = kpi_human_names.get(key, key.replace("_", " ").title())
            if isinstance(val, float):
                val_s = f"{val}{unit}"
                tgt_s = f"{tgt}{unit}"
                # Determine pass/fail
                if key == "churn_auc":
                    status = "✓ PASS" if val >= tgt else "✗ FAIL"
                elif key in ("demand_mape", "stockout_rate"):
                    status = "✓ PASS" if val <= tgt else "✗ FAIL"
                else:
                    status = "✓ PASS" if val >= tgt * 0.85 else "✗ FAIL"
            else:
                val_s, tgt_s, status = str(val), str(tgt), "—"
            table_data.append([label, val_s, tgt_s, status])

        tbl = Table(table_data, colWidths=[7.5 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm])
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor(f"#{_PRIMARY_HEX}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#FAFAFA")]),
                    ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#EEEEEE")),
                    ("FONTSIZE", (0, 1), (-1, -1), 9),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, cm))

        # ── Chart images ───────────────────────────────────────────────
        story.append(Paragraph("<b>Charts</b>", styles["Heading3"]))
        for img_bytes in chart_images[:3]:
            if not img_bytes:
                continue
            img_buffer = io.BytesIO(img_bytes)
            try:
                img = Image(img_buffer, width=16 * cm, height=8.5 * cm)
                story.append(img)
                story.append(Spacer(1, 0.5 * cm))
            except Exception as exc:
                logger.warning("Could not embed chart image: %s", exc)

        # ── Footer ─────────────────────────────────────────────────────
        story.append(Spacer(1, cm))
        story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#EEEEEE")))
        story.append(
            Paragraph(
                "NeuralRetail AMX-DS-2026-04 — CONFIDENTIAL — Do not distribute",
                ParagraphStyle("footer", parent=styles["Normal"], fontSize=7, textColor=rl_colors.HexColor("#AAAAAA")),
            )
        )

        doc.build(story)
        pdf_bytes = buffer.getvalue()

        if output_path is not None:
            Path(output_path).write_bytes(pdf_bytes)
            logger.info("PDF report saved to %s", output_path)

        return pdf_bytes

    # ------------------------------------------------------------------
    # CRM CSV export
    # ------------------------------------------------------------------

    @staticmethod
    def export_crm_csv(
        df: pd.DataFrame,
        output_path: str | Path | None = None,
    ) -> bytes:
        """Export CRM customer data as UTF-8 BOM CSV for Excel compatibility.

        The BOM (``\\ufeff``) ensures that Microsoft Excel correctly interprets
        UTF-8 encoded non-ASCII characters in customer names and product categories.

        Args:
            df: Customer CRM DataFrame. Expected columns: customer_id, persona,
                churn_proba, clv_tier, top_action_code. Extra columns are preserved.
            output_path: Optional file path to save to disk alongside bytes return.

        Returns:
            UTF-8 BOM encoded CSV bytes.
        """
        csv_bytes = df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")

        if output_path is not None:
            Path(output_path).write_bytes(csv_bytes)
            logger.info("CRM CSV saved to %s (rows=%d)", output_path, len(df))

        return csv_bytes
